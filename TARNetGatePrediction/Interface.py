#coding: utf-8
import BuildModelAndPredict
import openpyxl as xl
import numpy as np
import decimal


def PredictAStation(Days):
    '''
    预测一个站点
    暂时是苏浙皖的数据
    :return:
    '''
    FirstResultMat = []
    for columnIndex in range(17, 37):#从19:30 到 5：00
        #从EXCEL拿数据
        TestData = xl.load_workbook('OriginalData.xlsx')
        WorkSheet = TestData.get_sheet_by_name(TestData.get_sheet_names()[6])
        DataMat = []
        for index in range(6,84):
            DataMat.append(WorkSheet.cell(row=index, column= columnIndex).value)

        #从文件中拿训练数据
        TestDataArr = []
        for index in range(84, 91):
            TestDataArr.append(WorkSheet.cell(row=index, column=columnIndex).value)
        #训练找最好的L，D
        BestPredRes = []
        MinDiff = np.inf
        BestD = 0
        BestL = 0

        for D in range(1, 10):
            for L in range(1, 10):
                DataMatPara = DataMat[:]
                TempPredResult = BuildModelAndPredict.predict(DataMatPara, Days, D, L, )
                if None == TempPredResult:
                    continue
                SquareDiff = 0
                for index in range(0, Days):
                    if TempPredResult[index] < 0:
                        SquareDiff = np.inf
                        break
                    SquareDiff += np.abs(TestDataArr[index] - TempPredResult[index])
                if SquareDiff < MinDiff:
                    BestD = D; BestL = L
                    BestPredRes = TempPredResult
                    MinDiff = SquareDiff
        #用BestL, BestD 预测接下来的
        DataMat.extend(TestDataArr)#把训练数据也加入到拟合数据集中
        PredResult = BuildModelAndPredict.predict(DataMat, Days, BestD, BestL)
        FirstResultMat.append(PredResult)
        print columnIndex

    #把按照时间划分的预测结果转化成按照日期的并重新计算比例

    RealResMat = []
    for index in range(0,Days):
        DayPred = []
        DaySum = 0
        for index2 in range(0, 20):
            #如果预测的值是负数，就变为预测值中非零值的平均值
            if FirstResultMat[index2][index] < 0:
                ReplaceRate = 0
                Count = 0
                for i in range(0,Days):
                    if FirstResultMat[index2][i] > 0:
                        ReplaceRate += FirstResultMat[index2][i]
                        Count += 1
                FirstResultMat[index2][index] = ReplaceRate / Count
            #把结果加入到当天的列表中
            DayPred.append(FirstResultMat[index2][index])
            DaySum += FirstResultMat[index2][index]

        for index2 in range(0, 20):
            DayPred[index2] = 100 * DayPred[index2]  / DaySum

        RealResMat.append(DayPred)

    return RealResMat


Res = PredictAStation(7)

for index in range(0, len(Res)):
    print np.sum(Res[index])
    print Res[index]






